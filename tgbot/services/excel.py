from datetime import datetime
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
import os


class ExcelFile:
    def __init__(self):
        self.users_path = f"{os.getcwd()}/users_list.xlsx"
        self.files_path = f"{os.getcwd()}/files_list.xlsx"
        self.calculation_path = f"{os.getcwd()}/calculation_list.xlsx"

    @staticmethod
    def __reformat_date(date: Optional[datetime]) -> str:
        result_date = date.strftime("%d.%m.%Y") if date else "---"
        result_time = date.strftime("%H:%M") if date else "---"
        return result_date, result_time

    def create_users_file(self, users: List[dict]):
        wb = Workbook()
        ws = wb.active
        ws.append(
            (
                "User ID",
                "Username",
                "Дата регистрации (utc)",
                "Время регистрации (utc)",
                "Количество запросов"
            )
        )
        title_ft = Font(bold=True)
        for row in ws['A1:T1']:
            for cell in row:
                cell.font = title_ft

        for i, user in enumerate(users, start=2):
            ws.append(
                (
                    user["user_id"],
                    user["username"],
                    self.__reformat_date(user["create_dtime"])[0],
                    self.__reformat_date(user["create_dtime"])[1],
                    user["request_count"],
                )
            )
        wb.save(self.users_path)

    def create_images_file(self, files: List[dict]):
        wb = Workbook()
        ws = wb.active
        ws.append(
            (
                "File ID",
                "File_name",
            )
        )
        title_ft = Font(bold=True)
        for row in ws['A1:T1']:
            for cell in row:
                cell.font = title_ft

        for file in files:
            ws.append(
                (
                    file["file_id"],
                    file["file_name"],
                )
            )
        wb.save(self.files_path)

    def read_images_file(self):
        wb = load_workbook(filename=self.files_path)
        sh = wb.active
        result = []
        for row in sh.iter_rows(min_row=2):
            file_id = row[0].value
            file_name = row[1].value if row[1].value else ""
            result.append(dict(file_id=file_id, file_name=file_name))
        return result

    def create_calculation_file(self, net_to_seller: int, payments: List[dict]):
        wb = load_workbook(filename=f"calculation_template.xlsx")
        ws = wb.active
        # net_to_seller = f"AED {'{0:,}'.format(net_to_seller).replace(',', ' ')}"
        ws["C1"] = f"AED {'{0:,}'.format(net_to_seller).replace(',', ' ')}"
        bd = Side(style='thin', color="000000")
        ws.append(())
        total_payment = 0
        for i, item in enumerate(payments, start=2):
            one_payment_row = i + 7
            total_payment += item["payment_value"]
            ws.append(
                (
                    f"{i} payment",
                    item["payment_date"],
                    f"AED {'{0:,}'.format(item['payment_value']).replace(',', ' ')}",
                    # item["payment_value"],
                )
            )
            for cell in ws[f"A{one_payment_row}:C{one_payment_row}"][0]:
                cell.font = Font(name='Arial', size=11)
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            # ws[f"C{one_payment_row}"].number_format = "[$AED ]### ### ##0"

        if len(payments) > 0:
            post_payment_row = len(payments) + 9
            ws.append(
                (
                    "Total (post payment)",
                    "",
                    # total_payment,
                    f"AED {'{0:,}'.format(total_payment).replace(',', ' ')}",
                )
            )
            # ws[f"C{post_payment_row}"].number_format = "[$AED ]### ### ##0"
            ws.merge_cells(f"A{post_payment_row}:B{post_payment_row}")
            for cell in ws[f"A{post_payment_row}:C{post_payment_row}"][0]:
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            ws.append(())
            total_cost_row = len(payments) + 11
        else:
            total_cost_row = 9
        ws.append(
            (
                "Total cost for BUYER:	",
                "",
                f"AED {'{0:,}'.format(total_payment).replace(',', ' ')}",
                # total_payment,
            )
        )
        # ws[f"C{total_cost_row}"].number_format = "[$AED ]### ### ##0"
        for cell in ws[f"A{total_cost_row}:C{total_cost_row}"][0]:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell.fill = PatternFill(start_color="d9ead3", end_color="d9ead3", fill_type="solid")
        ws.merge_cells(f"A{total_cost_row}:B{total_cost_row}")
        initial_payment_to_seller = net_to_seller - total_payment
        initial_payment_to_seller_percent = (initial_payment_to_seller / net_to_seller)
        dld_transfer_fee = int((net_to_seller * 0.04) + 40)
        dld_registration_trusty_fee = 5250
        ww_agency_fee = int(net_to_seller * 0.02 * 1.05)
        total_on_transfer_date = initial_payment_to_seller + dld_transfer_fee + dld_registration_trusty_fee
        total_on_transfer_date += ww_agency_fee
        total_cost_for_buyer = total_payment + total_on_transfer_date
        # ws["C3"] = total_on_transfer_date
        ws["C3"] = f"AED {'{0:,}'.format(total_on_transfer_date).replace(',', ' ')}"
        # ws["C4"] = initial_payment_to_seller
        ws["C4"] = f"AED {'{0:,}'.format(initial_payment_to_seller).replace(',', ' ')}"
        ws["B4"] = initial_payment_to_seller_percent
        # ws["C5"] = dld_transfer_fee
        ws["C5"] = f"AED {'{0:,}'.format(dld_transfer_fee).replace(',', ' ')}"
        # ws["C6"] = dld_registration_trusty_fee
        ws["C6"] = f"AED {'{0:,}'.format(dld_registration_trusty_fee).replace(',', ' ')}"
        # ws["C7"] = ww_agency_fee
        ws["C7"] = f"AED {'{0:,}'.format(ww_agency_fee).replace(',', ' ')}"
        # ws[f"C{total_cost_row}"] = total_cost_for_buyer
        ws[f"C{total_cost_row}"] = f"AED {'{0:,}'.format(total_cost_for_buyer).replace(',', ' ')}"
        for row in ws["C:C"]:
            row.alignment = Alignment(horizontal='right')
        wb.save(self.calculation_path)


if __name__ == "__main__":
    excel = ExcelFile()
    excel.create_calculation_file(net_to_seller=2050000,
                                  payments=[{"payment_date": "22-09-2023", "payment_value": 159089}])
